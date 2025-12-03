#!/usr/bin/env python3
"""
Excel -> ASM Bitmap Generator for MARS/QtSPIM

Reads a color-painted Excel (.xlsx) worksheet and generates an assembly file that,
when included, defines a macro (default: drawMenu) which writes the corresponding
pixels into the MARS bitmap display starting at the address in $10. Each cell is
one pixel; each pixel is stored as a 32-bit word. Rows are emitted top-to-bottom,
left-to-right (row-major), writing contiguous words at offsets 0, 4, 8, ...

Palette mapping
- The script maps Excel cell colors to your pre-loaded palette registers, as set
  in main.asm's `colors:` routine (e.g., $20=0x000000, $21=0x4169E1, ...).
- If a cell color doesn't exactly match a palette color, the script picks the
  nearest palette color (Euclidean distance in RGB). Use --strict to require
  exact matches and fail otherwise.

Output style
- By default, generates a macro named drawMenu that expands to many `sw` using
  the palette registers and ends with `jr $31` (matching your current style).
- You can customize macro name or emit a label routine instead.

Usage (Windows cmd):
  py tools\\gen_bitmap_from_excel.py --input mymap.xlsx --sheet Sheet1 \
     --out sprites\\menuStart.asm --macro-name drawMenu

Requires: openpyxl (pip install openpyxl)
"""
import argparse
import os
import sys
from typing import Dict, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    print("This script requires openpyxl. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Palette mapping: Hex (no 0x) -> Constant Name
DEFAULT_PALETTE: Dict[str, str] = {
    '964B00': 'COLOR_BROWN',
    '00A8FF': 'COLOR_LIGHTBLUE',
    '00FF00': 'COLOR_GREEN',
    '000000': 'COLOR_BLACK',
    '4169E1': 'COLOR_ROYALBLUE',
    'FFFF00': 'COLOR_YELLOW',
    'CFBA95': 'COLOR_SKIN',
    'DC143C': 'COLOR_CRIMSON',
    '606060': 'COLOR_LIGHTGRAY',
    'FFA500': 'COLOR_ORANGE',
    'FF6600': 'COLOR_DARKORANGE',
    'FF0000': 'COLOR_RED',
    '808080': 'COLOR_GRAY',
    'FFFFFF': 'COLOR_WHITE',
}

RGBA_HEX_LEN = 8  # openpyxl returns ARGB as 8 hex digits (AARRGGBB)


def hex_to_rgb(hexstr: str) -> Tuple[int, int, int]:
    h = hexstr.strip().upper()
    # openpyxl often yields ARGB (AARRGGBB); strip alpha if present
    if len(h) == RGBA_HEX_LEN:
        h = h[2:]
    if h.startswith('0X'):
        h = h[2:]
    if len(h) != 6:
        raise ValueError(f"Invalid RGB hex '{hexstr}'")
    r = int(h[0:2], 16)
    g = int(h[2:4], 16)
    b = int(h[4:6], 16)
    return r, g, b


def rgb_to_hex(rgb: Tuple[int, int, int]) -> str:
    return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"


def nearest_palette(color_hex: str, palette: Dict[str, str]) -> Tuple[str, str]:
    """Return (palette_color_hex, palette_name) nearest to color_hex."""
    try:
        r, g, b = hex_to_rgb(color_hex)
    except Exception:
        return '000000', palette.get('000000', 'COLOR_BLACK')

    best_hex = None
    best_name = None
    best_d2 = None
    for phex, pname in palette.items():
        pr, pg, pb = hex_to_rgb(phex)
        dr = r - pr
        dg = g - pg
        db = b - pb
        d2 = dr*dr + dg*dg + db*db
        if best_d2 is None or d2 < best_d2:
            best_d2 = d2
            best_hex = phex
            best_name = pname
    return best_hex or '000000', best_name or 'COLOR_BLACK'


def read_sheet_colors(xlsx_path: str, sheet_name: str = None, max_rows: int = 64, max_cols: int = 128):
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print("Sheet not found: {}".format(sheet_name), file=sys.stderr)
            print("Available sheets: {}".format(
                ", ".join(wb.sheetnames)
            ), file=sys.stderr)
            sys.exit(3)
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    # Use fixed dimensions: start at row 1, col 1, and read max_rows x max_cols
    min_row = 1
    max_row = max_rows
    min_col = 1
    max_col = max_cols

    pixels = []  # list of palette register names per pixel in row-major order
    raw_colors = []  # raw hex collected for stats

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            fill = cell.fill
            rgb = None
            # Try different ways to get color
            if fill and getattr(fill, 'start_color', None):
                sc = fill.start_color
                if sc and sc.type == 'rgb' and sc.rgb:
                    rgb = sc.rgb
                elif sc and sc.type == 'indexed' and sc.indexed is not None:
                    # Indexed colors are Excel palette indices; map common ones or default to white
                    # We'll treat as white for simplicity
                    rgb = 'FFFFFFFF'
            if rgb is None and fill and getattr(fill, 'fgColor', None):
                fc = fill.fgColor
                if fc and fc.type == 'rgb' and fc.rgb:
                    rgb = fc.rgb
            if rgb is None:
                # Empty/no fill: default to white so it's visible
                rgb = 'FFFFFFFF'
            raw_colors.append(rgb)
            yield rgb


def generate_asm(
    color_groups: Dict[str, Dict],
    out_path: str,
    macro_name: str = 'drawMenu',
    comment: str = None,
):
    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    lines = []
    lines.append('# Auto-sprites by tools/gen_bitmap_from_excel.py')
    if comment:
        lines.append(f'# {comment}')
    lines.append('')
    
    # Emit .eqv definitions for used colors
    lines.append('# Color Palette Definitions')
    for hex_val, data in color_groups.items():
        lines.append(f'.eqv {data["name"]} 0x{hex_val}')
    
    lines.append('')
    lines.append('.text')
    lines.append(f'.macro {macro_name}')
    lines.append(f'{macro_name}:')
    
    # Generate optimized drawing code
    # Iterate through colors
    for hex_val, data in color_groups.items():
        name = data['name']
        offsets = data['offsets']
        
        if not offsets:
            continue
            
        lines.append(f'    # Draw {name}')
        lines.append(f'    li $t0, {name}')
        
        for offset in offsets:
            lines.append(f'    sw $t0, {offset}($10)')
            
    lines.append(f'.end_macro')

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines) + "\n")


def main():
    ap = argparse.ArgumentParser(description='Generate ASM bitmap macro from Excel colors')
    ap.add_argument('--input', required=True, help='Path to input .xlsx')
    ap.add_argument('--sheet', default=None, help='Worksheet name (default: active)')
    ap.add_argument('--out', required=True, help='Path to output .asm (e.g., sprites/menuStart.asm)')
    ap.add_argument('--macro-name', default='drawMenu', help='Macro name to generate (default: drawMenu)')
    ap.add_argument('--strict', action='store_true', help='Require exact palette matches; error on unknown colors')
    ap.add_argument('--rows', type=int, default=64, help='Number of rows to read (default: 64)')
    ap.add_argument('--cols', type=int, default=128, help='Number of columns to read (default: 128)')
    args = ap.parse_args()

    palette = DEFAULT_PALETTE

    # Resolve paths: allow input relative to current working directory OR script directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    cwd = os.getcwd()

    input_path = args.input
    if not os.path.isabs(input_path):
        # Try as given relative to CWD
        cand1 = os.path.abspath(os.path.join(cwd, input_path))
        # Try relative to script directory
        cand2 = os.path.abspath(os.path.join(script_dir, input_path))
        if os.path.isfile(cand1):
            input_path = cand1
        elif os.path.isfile(cand2):
            input_path = cand2
        else:
            print(f"Input file not found: {args.input}", file=sys.stderr)
            print(f"Tried: {cand1}\n       {cand2}", file=sys.stderr)
            print("Tip: If the file is next to this script in 'tools/', use --input tools/mymap.xlsx when running from repo root, or cd tools and pass --input mymap.xlsx", file=sys.stderr)
            sys.exit(4)
    else:
        if not os.path.isfile(input_path):
            print(f"Input file not found: {input_path}", file=sys.stderr)
            sys.exit(4)

    out_path = args.out
    if not os.path.isabs(out_path):
        # Leave relative to CWD as-is so users can write 'menuStart.asm' into repo root
        out_path = os.path.abspath(os.path.join(cwd, out_path))

    # Read cell fills
    raw_rgbs = list(read_sheet_colors(input_path, args.sheet, args.rows, args.cols))
    
    # Structure: { hex_string: { 'name': 'COLOR_NAME', 'offsets': [0, 4, 8...] } }
    color_groups = {}
    
    unknown_count = 0
    approximated_colors = {}
    
    offset = 0
    for rgb in raw_rgbs:
        hex6 = rgb[2:].upper() if len(rgb) == RGBA_HEX_LEN else rgb.upper()
        hex6 = hex6[-6:]
        
        target_hex = hex6
        target_name = ""
        
        if hex6 in palette:
            target_name = palette[hex6]
        else:
            if args.strict:
                print(f"Unknown color {hex6} not in palette; run without --strict to use nearest.", file=sys.stderr)
                sys.exit(2)
            
            nearest_hex, nearest_name = nearest_palette(hex6, palette)
            target_hex = nearest_hex
            target_name = nearest_name
            
            if hex6 not in approximated_colors:
                approximated_colors[hex6] = {
                    'count': 0,
                    'nearest': nearest_hex,
                    'name': nearest_name
                }
            approximated_colors[hex6]['count'] += 1
            unknown_count += 1
        
        # Add to group
        if target_hex not in color_groups:
            color_groups[target_hex] = {'name': target_name, 'offsets': []}
        
        color_groups[target_hex]['offsets'].append(offset)
        offset += 4

    # Generate ASM
    rel_comment = os.path.basename(input_path)
    generate_asm(
        color_groups=color_groups,
        out_path=out_path,
        macro_name=args.macro_name,
        comment=f"Source Excel: {rel_comment}; unknown colors matched: {unknown_count}",
    )

    print(f"Input : {input_path}")
    print(f"Output: {out_path}")
    print(f"Pixels: {len(raw_rgbs)}")
    
    if unknown_count:
        print(f"\n{'='*70}")
        print(f"Note: {unknown_count} pixels were approximated to nearest palette colors")
        print(f"{'='*70}")
        print(f"{'Excel Color':<15} {'Mapped To':<15} {'Name':<20} {'Count':<10}")
        print(f"{'-'*70}")
        
        # Sort by count (most frequent first)
        sorted_colors = sorted(approximated_colors.items(), 
                              key=lambda x: x[1]['count'], 
                              reverse=True)
        
        for original_hex, info in sorted_colors:
            print(f"0x{original_hex:<13} 0x{info['nearest']:<13} {info['name']:<20} {info['count']:<10}")
        print(f"{'='*70}")
        print(f"{'Excel Color':<15} {'Mapped To':<15} {'Register':<10} {'Count':<10}")

if __name__ == '__main__':
    main()
